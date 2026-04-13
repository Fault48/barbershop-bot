import logging
import sqlite3
import os
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    ContextTypes, ConversationHandler, MessageHandler, filters
)

# ─── НАСТРОЙКИ ───────────────────────────────────────────────
BOT_TOKEN = "8673066375:AAFVmJe_IolAMQnIBcjFksEogx3w_RA9vSk"
ADMIN_IDS = []  # сюда добавь свой Telegram ID (число), например: [123456789]

MASTERS = ["Мастер 1", "Мастер 2", "Мастер 3"]
SERVICES = {
    "Стрижка": 500,
    "Борода": 400,
    "Укладка": 300,
}
WORK_HOURS = list(range(10, 20))  # 10:00 – 19:00

# ─── СОСТОЯНИЯ ДИАЛОГА ───────────────────────────────────────
(SELECT_MASTER, SELECT_SERVICE, SELECT_DATE, SELECT_TIME,
 CONFIRM, CANCEL_SELECT) = range(6)

DB_PATH = "barbershop.db"

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ─── БАЗА ДАННЫХ ─────────────────────────────────────────────
def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS appointments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                username TEXT,
                full_name TEXT,
                master TEXT,
                service TEXT,
                price INTEGER,
                date TEXT,
                time TEXT,
                created_at TEXT,
                status TEXT DEFAULT 'active'
            )
        """)
        conn.commit()


def save_appointment(user_id, username, full_name, master, service, price, date, time):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            INSERT INTO appointments
            (user_id, username, full_name, master, service, price, date, time, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user_id, username, full_name, master, service, price, date, time,
              datetime.now().strftime("%Y-%m-%d %H:%M")))
        conn.commit()


def get_user_appointments(user_id):
    with sqlite3.connect(DB_PATH) as conn:
        return conn.execute("""
            SELECT id, master, service, date, time, price
            FROM appointments
            WHERE user_id=? AND status='active'
            ORDER BY date, time
        """, (user_id,)).fetchall()


def cancel_appointment(appt_id, user_id):
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            UPDATE appointments SET status='cancelled'
            WHERE id=? AND user_id=?
        """, (appt_id, user_id))
        conn.commit()


def is_slot_taken(master, date, time):
    with sqlite3.connect(DB_PATH) as conn:
        row = conn.execute("""
            SELECT id FROM appointments
            WHERE master=? AND date=? AND time=? AND status='active'
        """, (master, date, time)).fetchone()
        return row is not None


def get_all_appointments():
    with sqlite3.connect(DB_PATH) as conn:
        return conn.execute("""
            SELECT id, full_name, username, master, service, price, date, time, status, created_at
            FROM appointments ORDER BY date, time
        """).fetchall()


# ─── ВСПОМОГАТЕЛЬНЫЕ ─────────────────────────────────────────
def get_next_days(n=7):
    days = []
    today = datetime.now().date()
    for i in range(n):
        d = today + timedelta(days=i)
        days.append(d.strftime("%Y-%m-%d"))
    return days


def fmt_date(date_str):
    d = datetime.strptime(date_str, "%Y-%m-%d")
    months = ["янв", "фев", "мар", "апр", "май", "июн",
              "июл", "авг", "сен", "окт", "ноя", "дек"]
    days_ru = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    return f"{days_ru[d.weekday()]} {d.day} {months[d.month-1]}"


# ─── СТАРТ ───────────────────────────────────────────────────
async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    kb = [
        [InlineKeyboardButton("✂️ Записаться", callback_data="book")],
        [InlineKeyboardButton("📋 Мои записи", callback_data="my_appointments")],
        [InlineKeyboardButton("❌ Отменить запись", callback_data="cancel_start")],
    ]
    await update.message.reply_text(
        "👋 Добро пожаловать в барбершоп!\n\nВыберите действие:",
        reply_markup=InlineKeyboardMarkup(kb)
    )


async def menu_button(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "book":
        return await ask_master(update, ctx)
    elif query.data == "my_appointments":
        return await show_my_appointments(update, ctx)
    elif query.data == "cancel_start":
        return await cancel_start(update, ctx)


# ─── ЗАПИСЬ: ВЫБОР МАСТЕРА ───────────────────────────────────
async def ask_master(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    kb = [[InlineKeyboardButton(m, callback_data=f"master_{i}")]
          for i, m in enumerate(MASTERS)]
    kb.append([InlineKeyboardButton("🏠 Главное меню", callback_data="home")])
    text = "💈 Выберите мастера:"
    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb))
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(kb))
    return SELECT_MASTER


async def select_master(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "home":
        return await go_home(update, ctx)
    idx = int(query.data.split("_")[1])
    ctx.user_data["master"] = MASTERS[idx]

    kb = [[InlineKeyboardButton(f"{s} — {p}₽", callback_data=f"service_{s}")]
          for s, p in SERVICES.items()]
    kb.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_master")])
    await query.edit_message_text(
        f"✅ Мастер: {MASTERS[idx]}\n\n💼 Выберите услугу:",
        reply_markup=InlineKeyboardMarkup(kb)
    )
    return SELECT_SERVICE


async def select_service(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "back_master":
        return await ask_master(update, ctx)

    service = query.data.split("_", 1)[1]
    ctx.user_data["service"] = service
    ctx.user_data["price"] = SERVICES[service]

    days = get_next_days(7)
    kb = [[InlineKeyboardButton(fmt_date(d), callback_data=f"date_{d}")] for d in days]
    kb.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_service")])
    await query.edit_message_text(
        f"✅ Услуга: {service}\n\n📅 Выберите дату:",
        reply_markup=InlineKeyboardMarkup(kb)
    )
    return SELECT_DATE


async def select_date(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "back_service":
        # повторно покажем выбор услуги
        kb = [[InlineKeyboardButton(f"{s} — {p}₽", callback_data=f"service_{s}")]
              for s, p in SERVICES.items()]
        kb.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_master")])
        await query.edit_message_text("💼 Выберите услугу:", reply_markup=InlineKeyboardMarkup(kb))
        return SELECT_SERVICE

    date = query.data.split("_", 1)[1]
    ctx.user_data["date"] = date
    master = ctx.user_data["master"]

    free_slots = [h for h in WORK_HOURS if not is_slot_taken(master, date, f"{h:02d}:00")]
    if not free_slots:
        await query.edit_message_text(
            f"😔 На {fmt_date(date)} у {master} нет свободного времени.\n\nВыберите другой день.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data=f"back_date")]])
        )
        return SELECT_DATE

    kb = []
    row = []
    for h in free_slots:
        row.append(InlineKeyboardButton(f"{h:02d}:00", callback_data=f"time_{h:02d}:00"))
        if len(row) == 4:
            kb.append(row)
            row = []
    if row:
        kb.append(row)
    kb.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_date")])

    await query.edit_message_text(
        f"✅ Дата: {fmt_date(date)}\n\n🕐 Выберите время:",
        reply_markup=InlineKeyboardMarkup(kb)
    )
    return SELECT_TIME


async def select_time(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "back_date":
        days = get_next_days(7)
        kb = [[InlineKeyboardButton(fmt_date(d), callback_data=f"date_{d}")] for d in days]
        kb.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_service")])
        await query.edit_message_text("📅 Выберите дату:", reply_markup=InlineKeyboardMarkup(kb))
        return SELECT_DATE

    time = query.data.split("_", 1)[1]
    ctx.user_data["time"] = time
    d = ctx.user_data
    summary = (
        f"📋 *Подтвердите запись:*\n\n"
        f"👤 Мастер: {d['master']}\n"
        f"💼 Услуга: {d['service']}\n"
        f"💰 Цена: {d['price']}₽\n"
        f"📅 Дата: {fmt_date(d['date'])}\n"
        f"🕐 Время: {d['time']}"
    )
    kb = [
        [InlineKeyboardButton("✅ Подтвердить", callback_data="confirm")],
        [InlineKeyboardButton("❌ Отмена", callback_data="home")],
    ]
    await query.edit_message_text(summary, reply_markup=InlineKeyboardMarkup(kb), parse_mode="Markdown")
    return CONFIRM


async def confirm_booking(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "home":
        return await go_home(update, ctx)

    user = query.from_user
    d = ctx.user_data
    save_appointment(
        user.id, user.username or "", user.full_name,
        d["master"], d["service"], d["price"], d["date"], d["time"]
    )
    await query.edit_message_text(
        f"🎉 Запись подтверждена!\n\n"
        f"👤 {d['master']}\n"
        f"💼 {d['service']}\n"
        f"📅 {fmt_date(d['date'])} в {d['time']}\n\n"
        f"До встречи! ✂️"
    )
    return ConversationHandler.END


# ─── МОИ ЗАПИСИ ──────────────────────────────────────────────
async def show_my_appointments(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id
    appts = get_user_appointments(user_id)
    if not appts:
        await query.edit_message_text(
            "У вас нет активных записей.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Меню", callback_data="home")]])
        )
        return ConversationHandler.END

    text = "📋 *Ваши записи:*\n\n"
    for a in appts:
        text += f"• {fmt_date(a[3])} {a[4]} — {a[1]}, {a[2]} ({a[5]}₽)\n"

    await query.edit_message_text(
        text,
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Меню", callback_data="home")]])
    )
    return ConversationHandler.END


# ─── ОТМЕНА ЗАПИСИ ───────────────────────────────────────────
async def cancel_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id
    appts = get_user_appointments(user_id)
    if not appts:
        await query.edit_message_text(
            "У вас нет активных записей для отмены.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Меню", callback_data="home")]])
        )
        return ConversationHandler.END

    kb = []
    for a in appts:
        label = f"{fmt_date(a[3])} {a[4]} — {a[1]}"
        kb.append([InlineKeyboardButton(f"❌ {label}", callback_data=f"cancelid_{a[0]}")])
    kb.append([InlineKeyboardButton("⬅️ Назад", callback_data="home")])

    await query.edit_message_text(
        "Выберите запись для отмены:",
        reply_markup=InlineKeyboardMarkup(kb)
    )
    return CANCEL_SELECT


async def do_cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "home":
        return await go_home(update, ctx)
    appt_id = int(query.data.split("_")[1])
    cancel_appointment(appt_id, query.from_user.id)
    await query.edit_message_text(
        "✅ Запись отменена.",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Меню", callback_data="home")]])
    )
    return ConversationHandler.END


# ─── ЭКСПОРТ В EXCEL (для админа) ────────────────────────────
async def export_excel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("⛔ Нет доступа.")
        return

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    appts = get_all_appointments()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Записи"

    headers = ["ID", "Имя", "Username", "Мастер", "Услуга", "Цена", "Дата", "Время", "Статус", "Создано"]
    header_fill = PatternFill("solid", fgColor="2B547E")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, a in enumerate(appts, 2):
        for col, val in enumerate(a, 1):
            ws.cell(row=row, column=col, value=val)

    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15

    path = "/tmp/barbershop_export.xlsx"
    wb.save(path)

    with open(path, "rb") as f:
        await update.message.reply_document(
            document=f,
            filename=f"barbershop_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            caption="📊 Все записи барбершопа"
        )


# ─── ВСПОМОГАТЕЛЬНЫЕ ХЭНДЛЕРЫ ────────────────────────────────
async def go_home(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    kb = [
        [InlineKeyboardButton("✂️ Записаться", callback_data="book")],
        [InlineKeyboardButton("📋 Мои записи", callback_data="my_appointments")],
        [InlineKeyboardButton("❌ Отменить запись", callback_data="cancel_start")],
    ]
    await query.edit_message_text("Выберите действие:", reply_markup=InlineKeyboardMarkup(kb))
    return ConversationHandler.END


async def home_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    return await go_home(update, ctx)


# ─── ЗАПУСК ──────────────────────────────────────────────────
def main():
    init_db()
    app = Application.builder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(menu_button, pattern="^(book|my_appointments|cancel_start)$")
        ],
        states={
            SELECT_MASTER: [CallbackQueryHandler(select_master, pattern="^(master_|home)")],
            SELECT_SERVICE: [CallbackQueryHandler(select_service, pattern="^(service_|back_)")],
            SELECT_DATE: [CallbackQueryHandler(select_date, pattern="^(date_|back_)")],
            SELECT_TIME: [CallbackQueryHandler(select_time, pattern="^(time_|back_)")],
            CONFIRM: [CallbackQueryHandler(confirm_booking, pattern="^(confirm|home)$")],
            CANCEL_SELECT: [CallbackQueryHandler(do_cancel, pattern="^(cancelid_|home)$")],
        },
        fallbacks=[CallbackQueryHandler(home_callback, pattern="^home$")],
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("export", export_excel))
    app.add_handler(conv)
    app.add_handler(CallbackQueryHandler(home_callback, pattern="^home$"))

    logger.info("Бот запущен!")
    app.run_polling()


if __name__ == "__main__":
    main()
